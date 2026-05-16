"""Render confirmed.json (+ plan.json + findings/) into an .xlsx workbook.

Slice C of the code-audit pack. Optional companion to ``render_html.py``;
HTML rendering does not depend on this script. ``openpyxl`` is a hard
project dependency (declared in ``pyproject.toml``); we still guard the
import so the CLI surfaces a clean message when run in a stripped-down
environment that intentionally excludes it.

Workbook layout (per ``references/report-schema.md``):

- Sheet 1 "Findings" — one row per confirmed finding (severity color-coded)
- Sheet 2 "Summary"  — pivot: severity × module + Totals
- Sheet 3 "RunMeta"  — key/value: run_id, target, generated_at, pack_version,
                       total/by-severity counts
- Sheet 4 "Rejected" — rejected + needs_more_evidence findings (audit trail)

CLI usage::

    python -m render_xlsx --run-id audit-2026-05-16-0435

Library usage::

    from render_xlsx import render_workbook
    result = render_workbook(workspace_root=Path("."), run_id="audit-...")

Reuses validation logic from ``render_html.py`` so both renderers agree on
what a "valid finding" looks like; if the two were to diverge, mixed
HTML/xlsx output would silently drift.
"""

from __future__ import annotations

import argparse
import dataclasses
import datetime as dt
import json
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from render_html import (
    REQUIRED_EVIDENCE_FIELDS,
    REQUIRED_FINDING_FIELDS,
    REQUIRED_REVIEWER_FIELDS,
    REQUIRED_VERIFIER_FIELDS,
    VALID_CATEGORIES,
    VALID_CONFIDENCES,
    VALID_SEVERITIES,
    VALID_VERIFIER_STATUSES,
    ReportError,
    _validate_findings,
    _collect_rejected,
)

PACK_VERSION = "0.1.0"

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    _OPENPYXL_AVAILABLE = True
    _OPENPYXL_IMPORT_ERROR: str | None = None
except ImportError as exc:  # pragma: no cover - env-dependent
    Workbook = None  # type: ignore[assignment,misc]
    Alignment = None  # type: ignore[assignment,misc]
    Font = None  # type: ignore[assignment,misc]
    PatternFill = None  # type: ignore[assignment,misc]
    get_column_letter = None  # type: ignore[assignment]
    _OPENPYXL_AVAILABLE = False
    _OPENPYXL_IMPORT_ERROR = str(exc)


SEVERITY_FILL_COLORS = {
    "critical": "FFC0392B",
    "high":     "FFE67E22",
    "medium":   "FFF39C12",
    "low":      "FF27AE60",
    "info":     "FF7F8C8D",
}
HEADER_FILL_COLOR = "FF34495E"

FINDING_COLUMNS = [
    ("id", 22),
    ("module", 18),
    ("file", 50),
    ("line_start", 11),
    ("line_end", 11),
    ("title", 50),
    ("category", 18),
    ("severity", 12),
    ("confidence", 12),
    ("description", 60),
    ("code_snippet", 60),
    ("reasoning", 60),
    ("trigger_conditions", 40),
    ("expected_vs_actual", 40),
    ("suggested_fix", 50),
    ("reviewer_agent", 28),
    ("reviewer_ts", 22),
    ("verifier_status", 18),
    ("verifier_reason", 50),
    ("verifier_evidence_check", 50),
    ("severity_before", 14),
]
CODE_SNIPPET_MAX_CHARS = 500


@dataclasses.dataclass(frozen=True)
class XlsxResult:
    output_path: Path
    finding_count: int
    by_severity: dict[str, int]
    by_module: dict[str, int]
    rejected_count: int
    skipped: bool
    skipped_reason: str | None = None


def render_workbook(
    *,
    workspace_root: Path | None = None,
    run_id: str | None = None,
    confirmed_path: Path | None = None,
    plan_path: Path | None = None,
    findings_dir: Path | None = None,
    output_path: Path | None = None,
    strict: bool = False,
) -> XlsxResult:
    """Render an .xlsx report.

    Either pass (``workspace_root`` + ``run_id``) — paths derive per the
    ``.garage/code-audit/runs/<run_id>/`` layout — or pass explicit paths.

    When ``strict=False`` (default), a missing ``openpyxl`` installation
    returns an ``XlsxResult`` with ``skipped=True`` instead of raising,
    matching the SKILL.md contract ("HTML rendering does not depend on
    xlsx; an env missing openpyxl should not block report generation").

    When ``strict=True`` (used by callers that explicitly opted into xlsx),
    a missing ``openpyxl`` raises ``ReportError``.
    """
    confirmed_path, plan_path, findings_dir, output_path = _resolve_paths(
        workspace_root=workspace_root,
        run_id=run_id,
        confirmed_path=confirmed_path,
        plan_path=plan_path,
        findings_dir=findings_dir,
        output_path=output_path,
    )

    plan = _load_json(plan_path) if plan_path.is_file() else {}
    confirmed = _load_json(confirmed_path) or []
    if not isinstance(confirmed, list):
        raise ReportError(f"{confirmed_path} must contain a JSON array of findings")

    _validate_findings(confirmed)
    rejected_records = (
        _collect_rejected(findings_dir) if findings_dir and findings_dir.is_dir() else []
    )

    by_severity = Counter(f["severity"] for f in confirmed)
    by_module = Counter(f["module"] for f in confirmed)

    if not _OPENPYXL_AVAILABLE:
        reason = f"openpyxl not installed ({_OPENPYXL_IMPORT_ERROR})"
        if strict:
            raise ReportError(reason)
        return XlsxResult(
            output_path=output_path,
            finding_count=len(confirmed),
            by_severity=dict(by_severity),
            by_module=dict(by_module),
            rejected_count=len(rejected_records),
            skipped=True,
            skipped_reason=reason,
        )

    wb = Workbook()
    # Default sheet -> Findings.
    ws_findings = wb.active
    ws_findings.title = "Findings"
    _build_findings_sheet(ws_findings, confirmed)

    ws_summary = wb.create_sheet("Summary")
    _build_summary_sheet(ws_summary, plan, by_module, confirmed)

    ws_runmeta = wb.create_sheet("RunMeta")
    _build_runmeta_sheet(ws_runmeta, plan, confirmed, rejected_records)

    ws_rejected = wb.create_sheet("Rejected")
    _build_rejected_sheet(ws_rejected, rejected_records)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return XlsxResult(
        output_path=output_path,
        finding_count=len(confirmed),
        by_severity=dict(by_severity),
        by_module=dict(by_module),
        rejected_count=len(rejected_records),
        skipped=False,
    )


def _resolve_paths(
    *,
    workspace_root: Path | None,
    run_id: str | None,
    confirmed_path: Path | None,
    plan_path: Path | None,
    findings_dir: Path | None,
    output_path: Path | None,
) -> tuple[Path, Path, Path | None, Path]:
    if confirmed_path and output_path:
        return (
            confirmed_path,
            plan_path or (confirmed_path.parent / "plan.json"),
            findings_dir or (confirmed_path.parent / "findings"),
            output_path,
        )
    if workspace_root is None or run_id is None:
        raise ReportError(
            "Either (confirmed_path + output_path) or (workspace_root + run_id) must be provided"
        )
    run_dir = workspace_root / ".garage" / "code-audit" / "runs" / run_id
    return (
        run_dir / "confirmed.json",
        run_dir / "plan.json",
        run_dir / "findings",
        run_dir / "reports" / "report.xlsx",
    )


def _load_json(path: Path) -> Any:
    if not path.is_file():
        raise ReportError(f"Missing required file: {path}")
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ReportError(f"Invalid JSON in {path}: {exc}") from exc


def _truncate(text: str, max_chars: int) -> str:
    if len(text) <= max_chars:
        return text
    return text[: max_chars - 1] + "…"


def _build_findings_sheet(ws: Any, confirmed: list[dict[str, Any]]) -> None:
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    wrap = Alignment(wrap_text=True, vertical="top")

    # Header row + column widths.
    for col_idx, (name, width) in enumerate(FINDING_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"

    severity_fills = {
        sev: PatternFill("solid", fgColor=color)
        for sev, color in SEVERITY_FILL_COLORS.items()
    }
    severity_font_light = Font(color="FFFFFFFF", bold=True)

    for row_idx, f in enumerate(_sorted(confirmed), start=2):
        ev = f["evidence"]
        rv = f["reviewer"]
        vr = f["verifier"]
        values = [
            f["id"], f["module"], f["file"], f["line_start"], f["line_end"],
            f["title"], f["category"], f["severity"], f["confidence"],
            f["description"], _truncate(ev["code_snippet"], CODE_SNIPPET_MAX_CHARS),
            ev["reasoning"], ev["trigger_conditions"], ev["expected_vs_actual"],
            f["suggested_fix"], rv["agent"], rv["ts"],
            vr["status"], vr["reason"], vr["evidence_check"],
            f.get("severity_before", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap
        # Color the severity cell.
        severity_cell = ws.cell(row=row_idx, column=_column_index("severity"))
        severity_cell.fill = severity_fills.get(f["severity"], severity_fills["info"])
        severity_cell.font = severity_font_light


def _column_index(name: str) -> int:
    for i, (n, _) in enumerate(FINDING_COLUMNS, start=1):
        if n == name:
            return i
    raise KeyError(name)


def _sorted(findings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = {sev: i for i, sev in enumerate(VALID_SEVERITIES)}
    return sorted(
        findings,
        key=lambda f: (order.get(f["severity"], 99), f["module"], int(f["line_start"])),
    )


def _build_summary_sheet(
    ws: Any,
    plan: dict[str, Any],
    by_module: Counter[str],
    confirmed: list[dict[str, Any]],
) -> None:
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    bold = Font(bold=True)

    modules_in_plan = [m.get("name", "<unknown>") for m in plan.get("modules", [])]
    modules_seen = list(dict.fromkeys(modules_in_plan + sorted(by_module.keys())))
    if not modules_seen:
        ws.cell(row=1, column=1, value="(no findings to summarize)").font = bold
        return

    # Header: severity column then one column per module then Total.
    ws.cell(row=1, column=1, value="severity").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    for c_idx, m in enumerate(modules_seen, start=2):
        cell = ws.cell(row=1, column=c_idx, value=m)
        cell.font = header_font
        cell.fill = header_fill
    total_col = len(modules_seen) + 2
    cell = ws.cell(row=1, column=total_col, value="Total")
    cell.font = header_font
    cell.fill = header_fill

    severity_fills = {
        sev: PatternFill("solid", fgColor=color)
        for sev, color in SEVERITY_FILL_COLORS.items()
    }
    severity_font_light = Font(color="FFFFFFFF", bold=True)

    # Per-severity counts per module.
    by_module_severity: dict[str, Counter[str]] = {m: Counter() for m in modules_seen}
    for f in confirmed:
        by_module_severity.setdefault(f["module"], Counter())[f["severity"]] += 1

    for r_idx, sev in enumerate(VALID_SEVERITIES, start=2):
        sev_cell = ws.cell(row=r_idx, column=1, value=sev)
        sev_cell.fill = severity_fills[sev]
        sev_cell.font = severity_font_light
        row_total = 0
        for c_idx, m in enumerate(modules_seen, start=2):
            n = by_module_severity.get(m, Counter()).get(sev, 0)
            ws.cell(row=r_idx, column=c_idx, value=n)
            row_total += n
        total_cell = ws.cell(row=r_idx, column=total_col, value=row_total)
        total_cell.font = bold

    # Totals row.
    totals_row = len(VALID_SEVERITIES) + 2
    cell = ws.cell(row=totals_row, column=1, value="Total")
    cell.font = bold
    grand_total = 0
    for c_idx, m in enumerate(modules_seen, start=2):
        col_total = sum(by_module_severity.get(m, Counter()).values())
        cell = ws.cell(row=totals_row, column=c_idx, value=col_total)
        cell.font = bold
        grand_total += col_total
    cell = ws.cell(row=totals_row, column=total_col, value=grand_total)
    cell.font = bold

    # Column widths.
    ws.column_dimensions["A"].width = 14
    for c_idx in range(2, total_col + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 16


def _build_runmeta_sheet(
    ws: Any,
    plan: dict[str, Any],
    confirmed: list[dict[str, Any]],
    rejected: list[dict[str, Any]],
) -> None:
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)

    ws.cell(row=1, column=1, value="key").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=2, value="value").font = header_font
    ws.cell(row=1, column=2).fill = header_fill

    by_severity = Counter(f["severity"] for f in confirmed)
    rows = [
        ("run_id", str(plan.get("run_id") or "")),
        ("target", str(plan.get("target") or "")),
        ("generated_at", dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")),
        ("pack_version", PACK_VERSION),
        ("total_findings", len(confirmed)),
        ("modules_audited", len(plan.get("modules", []))),
        ("rejected_count", len(rejected)),
    ]
    for sev in VALID_SEVERITIES:
        rows.append((f"by_severity.{sev}", by_severity.get(sev, 0)))

    for r_idx, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r_idx, column=2, value=v)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50


def _build_rejected_sheet(ws: Any, rejected: list[dict[str, Any]]) -> None:
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor=HEADER_FILL_COLOR)
    wrap = Alignment(wrap_text=True, vertical="top")
    columns = [
        ("id", 22), ("module", 18), ("file", 50),
        ("status", 18), ("verifier_reason", 60),
    ]
    for c_idx, (name, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(c_idx)].width = width
    ws.freeze_panes = "A2"
    for r_idx, f in enumerate(rejected, start=2):
        vr = f.get("verifier") or {}
        values = [
            str(f.get("id", "")),
            str(f.get("module", "")),
            str(f.get("file", "")),
            str(vr.get("status", "")),
            str(vr.get("reason", "")),
        ]
        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = wrap


def _cli(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(
        prog="render_xlsx",
        description="Render code-audit confirmed.json into an .xlsx workbook.",
    )
    p.add_argument("--workspace", type=Path, default=Path.cwd(),
                   help="Workspace root (default: cwd). Used with --run-id.")
    p.add_argument("--run-id", type=str, default=None,
                   help="Run ID under .garage/code-audit/runs/<run-id>/.")
    p.add_argument("--confirmed-path", type=Path, default=None,
                   help="Explicit path to confirmed.json (overrides --workspace/--run-id).")
    p.add_argument("--plan-path", type=Path, default=None,
                   help="Explicit path to plan.json.")
    p.add_argument("--findings-dir", type=Path, default=None,
                   help="Explicit path to findings/ directory.")
    p.add_argument("--output", type=Path, default=None,
                   help="Explicit output xlsx path.")
    p.add_argument("--strict", action="store_true",
                   help="Treat missing openpyxl as a hard error (default: warn + skip).")
    args = p.parse_args(argv)

    try:
        result = render_workbook(
            workspace_root=args.workspace,
            run_id=args.run_id,
            confirmed_path=args.confirmed_path,
            plan_path=args.plan_path,
            findings_dir=args.findings_dir,
            output_path=args.output,
            strict=args.strict,
        )
    except ReportError as exc:
        print(f"render_xlsx: error: {exc}", file=sys.stderr)
        return 2

    if result.skipped:
        print(
            f"render_xlsx: skipped ({result.skipped_reason}); "
            "HTML report is unaffected. Install openpyxl to enable .xlsx output.",
            file=sys.stderr,
        )
        return 0

    print(
        f"Wrote {result.output_path} ({result.finding_count} findings, "
        f"{result.rejected_count} rejected/needs_more_evidence)",
        file=sys.stderr,
    )
    return 0


if __name__ == "__main__":  # pragma: no cover
    raise SystemExit(_cli())
