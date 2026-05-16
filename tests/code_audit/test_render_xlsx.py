"""Tests for packs/code-audit/skills/audit-reporter/scripts/render_xlsx.py.

Slice C of the code-audit pack. Verifies:

- Smoke: minimal valid input produces an .xlsx with the expected 4 sheets
- Findings sheet has 1 row per confirmed finding with all columns populated
- Summary sheet has pivot (severity × module) with correct totals
- RunMeta sheet has run_id / target / pack_version / counts
- Rejected sheet surfaces rejected and needs_more_evidence findings
- Severity color fills land on the correct cells (visual contract from
  ``references/report-schema.md``)
- Empty findings still produces a workbook (gracefully degrades)
- Schema validation (shared with render_html) catches bad input
- Missing-openpyxl path returns skipped=True in lenient mode, raises in
  strict mode (per SKILL.md "Excel rendering failure does not block HTML")
- CLI exit codes
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any
from unittest.mock import patch

import pytest
from openpyxl import load_workbook

import render_xlsx
from render_html import ReportError
from render_xlsx import (
    FINDING_COLUMNS,
    SEVERITY_FILL_COLORS,
    XlsxResult,
    render_workbook,
)


def _baseline_plan(run_id: str = "audit-test-001") -> dict[str, Any]:
    return {
        "schema_version": 1,
        "run_id": run_id,
        "target": "src/",
        "created_at": "2026-05-16T04:35:00Z",
        "budgets": {"module_budget_tokens": 30000, "module_budget_files": 20},
        "modules": [
            {
                "name": "runtime",
                "path": "src/runtime/",
                "priority": "high",
                "file_count": 3,
                "loc_estimate": 240,
                "languages": ["python"],
                "status": "done",
            },
            {
                "name": "adapter",
                "path": "src/adapter/",
                "priority": "medium",
                "file_count": 2,
                "loc_estimate": 180,
                "languages": ["python"],
                "status": "done",
            },
        ],
        "total_files": 5,
        "total_loc": 420,
    }


def _baseline_finding(
    *,
    fid: str = "F-001",
    severity: str = "high",
    category: str = "error-handling",
    confidence: str = "high",
    module: str = "runtime",
    file: str = "src/runtime/session_manager.py",
    line_start: int = 142,
    line_end: int = 148,
    file_sha256: str = "a" * 64,
    verifier_status: str = "confirmed",
    severity_before: str | None = None,
    severity_after: str | None = None,
) -> dict[str, Any]:
    f: dict[str, Any] = {
        "id": fid,
        "run_id": "audit-test-001",
        "module": module,
        "file": file,
        "line_start": line_start,
        "line_end": line_end,
        "file_sha256": file_sha256,
        "title": "KeyError 未被捕获",
        "category": category,
        "severity": severity,
        "confidence": confidence,
        "description": "self.sessions[session_id] 并发场景下抛 KeyError。",
        "evidence": {
            "code_snippet": "def _trigger():\n    meta = self.sessions[session_id]",
            "reasoning": "无 KeyError 防护，并发 archive 会触发。",
            "trigger_conditions": "并发 archive",
            "expected_vs_actual": "expected early-return; actual: KeyError",
            "related_files": ["src/runtime/session_manager.py:88"],
        },
        "suggested_fix": "self.sessions.get(session_id) + early-return.",
        "reviewer": {
            "agent": "code-audit-reviewer-agent",
            "ts": "2026-05-16T04:35:12Z",
        },
        "verifier": {
            "status": verifier_status,
            "reason": "已读源文件 L142-148。证据成立。",
            "evidence_check": "Read session_manager.py L88-148.",
            "agent": "code-audit-verifier-agent",
            "ts": "2026-05-16T04:42:55Z",
        },
    }
    if severity_before is not None:
        f["severity_before"] = severity_before
    if severity_after is not None:
        f["verifier"]["severity_after"] = severity_after
    return f


def _write_run(
    tmp_path: Path,
    *,
    run_id: str = "audit-test-001",
    confirmed: list[dict[str, Any]] | None = None,
    rejected_findings: list[dict[str, Any]] | None = None,
    plan: dict[str, Any] | None = None,
) -> Path:
    run_dir = tmp_path / ".garage" / "code-audit" / "runs" / run_id
    (run_dir / "findings").mkdir(parents=True)
    (run_dir / "reports").mkdir()
    if plan is None:
        plan = _baseline_plan(run_id=run_id)
    (run_dir / "plan.json").write_text(json.dumps(plan, indent=2), encoding="utf-8")
    if confirmed is None:
        confirmed = [_baseline_finding()]
    (run_dir / "confirmed.json").write_text(
        json.dumps(confirmed, indent=2), encoding="utf-8"
    )
    if rejected_findings:
        by_module: dict[str, list[dict[str, Any]]] = {}
        for f in rejected_findings:
            by_module.setdefault(f["module"], []).append(f)
        for module, items in by_module.items():
            (run_dir / "findings" / f"{module}.json").write_text(
                json.dumps(items, indent=2), encoding="utf-8"
            )
    return run_dir


class TestSmoke:
    def test_render_produces_xlsx(self, tmp_path: Path) -> None:
        _write_run(tmp_path)
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        assert isinstance(result, XlsxResult)
        assert result.output_path.is_file()
        assert result.skipped is False
        # File is a real xlsx (zip-archive signature 'PK')
        with open(result.output_path, "rb") as f:
            assert f.read(2) == b"PK"

    def test_workbook_has_four_sheets(self, tmp_path: Path) -> None:
        _write_run(tmp_path)
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        wb = load_workbook(result.output_path)
        assert wb.sheetnames == ["Findings", "Summary", "RunMeta", "Rejected"]


class TestFindingsSheet:
    def test_findings_sheet_has_header_and_rows(self, tmp_path: Path) -> None:
        findings = [
            _baseline_finding(fid="F-1", severity="critical"),
            _baseline_finding(fid="F-2", severity="high"),
            _baseline_finding(fid="F-3", severity="medium", module="adapter"),
        ]
        _write_run(tmp_path, confirmed=findings)
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        wb = load_workbook(result.output_path)
        ws = wb["Findings"]
        # Header row matches FINDING_COLUMNS exactly
        header = [ws.cell(row=1, column=c).value for c in range(1, len(FINDING_COLUMNS) + 1)]
        expected_header = [name for name, _ in FINDING_COLUMNS]
        assert header == expected_header
        # 3 data rows
        # Row sort: critical -> high -> medium per VALID_SEVERITIES order
        assert ws.cell(row=2, column=1).value == "F-1"  # critical first
        assert ws.cell(row=3, column=1).value == "F-2"  # high second
        assert ws.cell(row=4, column=1).value == "F-3"  # medium third

    def test_findings_evidence_columns_populated(self, tmp_path: Path) -> None:
        _write_run(tmp_path)
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        wb = load_workbook(result.output_path)
        ws = wb["Findings"]
        # Find column indices by header name.
        cols = {name: i + 1 for i, (name, _) in enumerate(FINDING_COLUMNS)}
        row = 2
        assert ws.cell(row=row, column=cols["code_snippet"]).value.startswith("def _trigger():")
        assert ws.cell(row=row, column=cols["reasoning"]).value.startswith("无 KeyError 防护")
        assert "expected" in ws.cell(row=row, column=cols["expected_vs_actual"]).value
        assert ws.cell(row=row, column=cols["verifier_status"]).value == "confirmed"
        assert ws.cell(row=row, column=cols["reviewer_agent"]).value == "code-audit-reviewer-agent"

    def test_severity_cell_has_color_fill(self, tmp_path: Path) -> None:
        findings = [_baseline_finding(fid=f"F-{i}", severity=sev)
                    for i, sev in enumerate(["critical", "high", "medium", "low", "info"], start=1)]
        _write_run(tmp_path, confirmed=findings)
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        wb = load_workbook(result.output_path)
        ws = wb["Findings"]
        sev_col = next(i + 1 for i, (n, _) in enumerate(FINDING_COLUMNS) if n == "severity")
        # Row 2..6 in critical/high/medium/low/info order.
        for r, sev in enumerate(["critical", "high", "medium", "low", "info"], start=2):
            cell = ws.cell(row=r, column=sev_col)
            assert cell.value == sev
            fill_color = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            expected = SEVERITY_FILL_COLORS[sev]
            assert fill_color == expected, (
                f"severity={sev}: got fill rgb={fill_color!r}, expected {expected!r}"
            )

    def test_findings_freeze_panes(self, tmp_path: Path) -> None:
        _write_run(tmp_path)
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        wb = load_workbook(result.output_path)
        ws = wb["Findings"]
        assert ws.freeze_panes == "A2"

    def test_code_snippet_truncated_at_500(self, tmp_path: Path) -> None:
        big_snippet = "x = 1\n" * 500  # > 500 chars
        finding = _baseline_finding()
        finding["evidence"]["code_snippet"] = big_snippet
        _write_run(tmp_path, confirmed=[finding])
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        wb = load_workbook(result.output_path)
        ws = wb["Findings"]
        cols = {name: i + 1 for i, (name, _) in enumerate(FINDING_COLUMNS)}
        snippet_value = ws.cell(row=2, column=cols["code_snippet"]).value
        assert snippet_value.endswith("…")
        assert len(snippet_value) <= 500


class TestSummarySheet:
    def test_summary_pivot_severity_by_module(self, tmp_path: Path) -> None:
        findings = [
            _baseline_finding(fid="F-1", severity="critical", module="runtime"),
            _baseline_finding(fid="F-2", severity="high",     module="runtime"),
            _baseline_finding(fid="F-3", severity="high",     module="adapter"),
            _baseline_finding(fid="F-4", severity="medium",   module="adapter"),
        ]
        _write_run(tmp_path, confirmed=findings)
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        wb = load_workbook(result.output_path)
        ws = wb["Summary"]
        # Header row: severity | runtime | adapter | Total
        assert ws.cell(row=1, column=1).value == "severity"
        # Modules columns may be ordered by plan order (runtime, adapter).
        module_cols = [ws.cell(row=1, column=c).value for c in range(2, 4)]
        assert set(module_cols) == {"runtime", "adapter"}
        assert ws.cell(row=1, column=4).value == "Total"

        # Find module column indices
        runtime_col = module_cols.index("runtime") + 2
        adapter_col = module_cols.index("adapter") + 2

        # critical row: runtime=1, adapter=0, total=1
        crit_row = next(
            r for r in range(2, 10) if ws.cell(row=r, column=1).value == "critical"
        )
        assert ws.cell(row=crit_row, column=runtime_col).value == 1
        assert ws.cell(row=crit_row, column=adapter_col).value == 0
        assert ws.cell(row=crit_row, column=4).value == 1

        # high row: runtime=1, adapter=1, total=2
        high_row = next(
            r for r in range(2, 10) if ws.cell(row=r, column=1).value == "high"
        )
        assert ws.cell(row=high_row, column=runtime_col).value == 1
        assert ws.cell(row=high_row, column=adapter_col).value == 1
        assert ws.cell(row=high_row, column=4).value == 2

        # medium row: runtime=0, adapter=1, total=1
        med_row = next(
            r for r in range(2, 10) if ws.cell(row=r, column=1).value == "medium"
        )
        assert ws.cell(row=med_row, column=runtime_col).value == 0
        assert ws.cell(row=med_row, column=adapter_col).value == 1
        assert ws.cell(row=med_row, column=4).value == 1

        # Totals row: grand total = 4
        total_row = next(
            r for r in range(2, 12) if ws.cell(row=r, column=1).value == "Total"
        )
        assert ws.cell(row=total_row, column=4).value == 4

    def test_summary_empty_findings_handled(self, tmp_path: Path) -> None:
        _write_run(tmp_path, confirmed=[])
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        wb = load_workbook(result.output_path)
        ws = wb["Summary"]
        # Either has the pivot with zeros, or has the empty placeholder.
        # Both shapes are acceptable; just make sure we didn't crash.
        first_cell = ws.cell(row=1, column=1).value
        assert first_cell in {"severity", "(no findings to summarize)"}


class TestRunMetaSheet:
    def test_runmeta_keys_and_values(self, tmp_path: Path) -> None:
        findings = [
            _baseline_finding(fid="F-1", severity="critical"),
            _baseline_finding(fid="F-2", severity="high"),
        ]
        _write_run(tmp_path, confirmed=findings)
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        wb = load_workbook(result.output_path)
        ws = wb["RunMeta"]
        keys = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                for r in range(2, ws.max_row + 1)}
        assert keys["run_id"] == "audit-test-001"
        assert keys["target"] == "src/"
        assert keys["pack_version"] == "0.1.0"
        assert keys["total_findings"] == 2
        assert keys["modules_audited"] == 2  # from plan
        assert keys["by_severity.critical"] == 1
        assert keys["by_severity.high"] == 1
        assert keys["by_severity.medium"] == 0
        assert "generated_at" in keys


class TestRejectedSheet:
    def test_rejected_findings_in_sheet(self, tmp_path: Path) -> None:
        confirmed = [_baseline_finding(fid="F-1")]
        rejected = [
            _baseline_finding(fid="F-R-1", verifier_status="rejected"),
            _baseline_finding(fid="F-R-2", verifier_status="needs_more_evidence"),
        ]
        _write_run(tmp_path, confirmed=confirmed, rejected_findings=rejected)
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        assert result.rejected_count == 2
        wb = load_workbook(result.output_path)
        ws = wb["Rejected"]
        assert ws.cell(row=1, column=1).value == "id"
        ids = {ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)}
        assert ids == {"F-R-1", "F-R-2"}
        statuses = {ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)}
        assert statuses == {"rejected", "needs_more_evidence"}


class TestSchemaValidation:
    def test_invalid_severity_raises(self, tmp_path: Path) -> None:
        bad = _baseline_finding()
        bad["severity"] = "URGENT"
        _write_run(tmp_path, confirmed=[bad])
        with pytest.raises(ReportError, match="invalid severity"):
            render_workbook(workspace_root=tmp_path, run_id="audit-test-001")

    def test_missing_confirmed_raises(self, tmp_path: Path) -> None:
        run_dir = tmp_path / ".garage" / "code-audit" / "runs" / "audit-x"
        run_dir.mkdir(parents=True)
        (run_dir / "plan.json").write_text("{}", encoding="utf-8")
        with pytest.raises(ReportError, match="Missing required file"):
            render_workbook(workspace_root=tmp_path, run_id="audit-x")

    def test_invalid_confirmed_json_raises(self, tmp_path: Path) -> None:
        run_dir = tmp_path / ".garage" / "code-audit" / "runs" / "audit-x"
        (run_dir / "findings").mkdir(parents=True)
        (run_dir / "reports").mkdir()
        (run_dir / "plan.json").write_text("{}", encoding="utf-8")
        (run_dir / "confirmed.json").write_text("not json {", encoding="utf-8")
        with pytest.raises(ReportError, match="Invalid JSON"):
            render_workbook(workspace_root=tmp_path, run_id="audit-x")


class TestOpenpyxlMissing:
    """SKILL.md contract: missing openpyxl -> skipped=True (lenient) or
    ReportError (strict)."""

    def test_missing_openpyxl_lenient_skips(self, tmp_path: Path) -> None:
        _write_run(tmp_path)
        with patch.object(render_xlsx, "_OPENPYXL_AVAILABLE", False), \
             patch.object(render_xlsx, "_OPENPYXL_IMPORT_ERROR", "test fake import error"):
            result = render_workbook(
                workspace_root=tmp_path,
                run_id="audit-test-001",
                strict=False,
            )
        assert result.skipped is True
        assert "openpyxl" in (result.skipped_reason or "")
        # Output xlsx should NOT have been written.
        assert not result.output_path.is_file()

    def test_missing_openpyxl_strict_raises(self, tmp_path: Path) -> None:
        _write_run(tmp_path)
        with patch.object(render_xlsx, "_OPENPYXL_AVAILABLE", False), \
             patch.object(render_xlsx, "_OPENPYXL_IMPORT_ERROR", "test fake import error"):
            with pytest.raises(ReportError, match="openpyxl"):
                render_workbook(
                    workspace_root=tmp_path,
                    run_id="audit-test-001",
                    strict=True,
                )


class TestExplicitPaths:
    def test_render_with_explicit_paths(self, tmp_path: Path) -> None:
        run_dir = _write_run(tmp_path)
        out = tmp_path / "custom-report.xlsx"
        result = render_workbook(
            confirmed_path=run_dir / "confirmed.json",
            plan_path=run_dir / "plan.json",
            findings_dir=run_dir / "findings",
            output_path=out,
        )
        assert out.is_file()
        assert result.output_path == out

    def test_render_without_run_id_or_paths_raises(self, tmp_path: Path) -> None:
        with pytest.raises(ReportError, match="Either"):
            render_workbook(workspace_root=tmp_path)


class TestSeverityUpgradeColumn:
    def test_severity_before_after_recorded(self, tmp_path: Path) -> None:
        finding = _baseline_finding(
            severity="high",
            severity_before="medium",
            severity_after="high",
            verifier_status="upgrade",
        )
        _write_run(tmp_path, confirmed=[finding])
        result = render_workbook(workspace_root=tmp_path, run_id="audit-test-001")
        wb = load_workbook(result.output_path)
        ws = wb["Findings"]
        cols = {name: i + 1 for i, (name, _) in enumerate(FINDING_COLUMNS)}
        assert ws.cell(row=2, column=cols["severity_before"]).value == "medium"
        assert ws.cell(row=2, column=cols["verifier_status"]).value == "upgrade"


class TestCliEntry:
    def test_cli_writes_xlsx(self, tmp_path: Path) -> None:
        _write_run(tmp_path)
        rc = render_xlsx._cli([
            "--workspace", str(tmp_path),
            "--run-id", "audit-test-001",
        ])
        assert rc == 0
        out = (
            tmp_path / ".garage" / "code-audit" / "runs"
            / "audit-test-001" / "reports" / "report.xlsx"
        )
        assert out.is_file()

    def test_cli_invalid_data_returns_2(self, tmp_path: Path) -> None:
        run_dir = tmp_path / ".garage" / "code-audit" / "runs" / "audit-x"
        run_dir.mkdir(parents=True)
        (run_dir / "plan.json").write_text("{}", encoding="utf-8")
        (run_dir / "confirmed.json").write_text(
            json.dumps([{"id": "bad"}]), encoding="utf-8"
        )
        rc = render_xlsx._cli([
            "--workspace", str(tmp_path),
            "--run-id", "audit-x",
        ])
        assert rc == 2

    def test_cli_missing_openpyxl_lenient(self, tmp_path: Path, capsys) -> None:
        _write_run(tmp_path)
        with patch.object(render_xlsx, "_OPENPYXL_AVAILABLE", False), \
             patch.object(render_xlsx, "_OPENPYXL_IMPORT_ERROR", "test fake import error"):
            rc = render_xlsx._cli([
                "--workspace", str(tmp_path),
                "--run-id", "audit-test-001",
            ])
        assert rc == 0  # lenient = exit 0 with stderr warning
        captured = capsys.readouterr()
        assert "skipped" in captured.err

    def test_cli_missing_openpyxl_strict(self, tmp_path: Path) -> None:
        _write_run(tmp_path)
        with patch.object(render_xlsx, "_OPENPYXL_AVAILABLE", False), \
             patch.object(render_xlsx, "_OPENPYXL_IMPORT_ERROR", "test fake import error"):
            rc = render_xlsx._cli([
                "--workspace", str(tmp_path),
                "--run-id", "audit-test-001",
                "--strict",
            ])
        assert rc == 2  # strict = error
